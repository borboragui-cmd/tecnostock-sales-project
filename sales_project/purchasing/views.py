from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from django.db.models import F, Avg, Sum, Count
from django.db.models.deletion import ProtectedError
from django.core.exceptions import ValidationError
from creditos_compras import services as creditos_compras_services

from decimal import Decimal
from io import BytesIO
from datetime import datetime
import html as html_module

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Purchase, PurchaseDetail
from .forms import PurchaseForm, PurchaseDetailFormSet
from billing.models import Product, Supplier


@login_required
def purchase_list(request):
    purchases = Purchase.objects.select_related('supplier').all()

    # Filtro por proveedor
    supplier_id = request.GET.get('supplier', '')
    if supplier_id:
        purchases = purchases.filter(supplier_id=supplier_id)

    # Filtro por rango de fechas (__range / __gte / __lte)
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from and date_to:
        purchases = purchases.filter(purchase_date__date__range=[date_from, date_to])
    elif date_from:
        purchases = purchases.filter(purchase_date__date__gte=date_from)
    elif date_to:
        purchases = purchases.filter(purchase_date__date__lte=date_to)

    # Filtro por año (__year)
    year = request.GET.get('year', '')
    if year:
        purchases = purchases.filter(purchase_date__year=year)

    suppliers = Supplier.objects.filter(is_active=True).order_by('name')

    return render(request, 'purchasing/purchase_list.html', {
        'purchases': purchases,
        'suppliers': suppliers,
        'supplier_id': supplier_id,
        'date_from': date_from,
        'date_to': date_to,
        'year': year,
    })


@login_required
def purchase_create(request):
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        formset = PurchaseDetailFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    purchase = form.save(commit=False)
                    purchase.save()
                    formset.instance = purchase
                    formset.save()

                    details = purchase.details.all()
                    subtotal = sum(d.subtotal for d in details)
                    purchase.subtotal = subtotal
                    purchase.tax = subtotal * Decimal('0.15')
                    purchase.total = purchase.subtotal + purchase.tax
                    purchase.save()

                    for detail in details:
                        Product.objects.filter(pk=detail.product.pk).update(
                            stock=F('stock') + detail.quantity
                        )

                    creditos_compras_services.procesar_tipo_pago(
                        purchase,
                        form.cleaned_data['tipo_pago'],
                        form.cleaned_data.get('num_cuotas'),
                    )
            except ValidationError as e:
                messages.error(request, str(e))
            else:
                messages.success(request, f'Compra #{purchase.numero} registrada. Total: ${purchase.total}')
                return redirect('purchasing:purchase_list')
    else:
        form = PurchaseForm()
        formset = PurchaseDetailFormSet()

    import json
    products_prices = json.dumps(
        {str(p.id): str(p.unit_price) for p in Product.objects.filter(is_active=True)}
    )
    return render(request, 'purchasing/purchase_form.html', {
        'form': form,
        'formset': formset,
        'title': 'Nueva Compra',
        'products_prices': products_prices,
    })


@login_required
def purchase_detail(request, pk):
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier').prefetch_related('details__product'),
        pk=pk
    )
    return render(request, 'purchasing/purchase_detail.html', {'purchase': purchase})


@login_required
def purchase_delete(request, pk):
    purchase = get_object_or_404(
        Purchase.objects.prefetch_related('details__product'),
        pk=pk,
    )
    if request.method == 'POST':
        purchase_id = purchase.id
        purchase_numero = purchase.numero
        try:
            with transaction.atomic():
                for detail in purchase.details.all():
                    Product.objects.filter(pk=detail.product.pk).update(
                        stock=F('stock') - detail.quantity
                    )
                purchase.delete()
        except ProtectedError:
            messages.error(
                request,
                f'No se puede eliminar la Compra #{purchase_numero}: tiene cuotas de '
                f'crédito asociadas. Elimina o reasigna las cuotas primero.'
            )
            return redirect('purchasing:purchase_list')
        messages.success(request, f'Compra #{purchase_id} eliminada y stock revertido.')
        return redirect('purchasing:purchase_list')
    return render(request, 'purchasing/purchase_confirm_delete.html', {'object': purchase})


@login_required
def purchase_pdf(request, pk):
    """Genera y descarga un PDF de la orden de compra."""
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier').prefetch_related('details__product'),
        pk=pk,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40,
    )
    elements = []
    styles = getSampleStyleSheet()
    esc = html_module.escape
    available_width = letter[0] - 80

    # ── Estilos ──────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        'PurchTitle', parent=styles['Heading1'],
        fontSize=22, textColor=colors.HexColor('#0F172A'),
        alignment=1, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'PurchSub', parent=styles['Normal'],
        fontSize=11, textColor=colors.grey, alignment=1, spaceAfter=16,
    )
    label_style = ParagraphStyle('PurchLabel', parent=styles['Normal'],
                                 fontSize=10, fontName='Helvetica-Bold')
    value_style = ParagraphStyle('PurchValue', parent=styles['Normal'], fontSize=10)
    hdr_style   = ParagraphStyle('TblHdr', parent=styles['Normal'],
                                 fontSize=10, fontName='Helvetica-Bold',
                                 textColor=colors.whitesmoke)
    cell_style  = ParagraphStyle('TblCell', parent=styles['Normal'], fontSize=10, leading=14)
    cell_r      = ParagraphStyle('TblCellR', parent=styles['Normal'], fontSize=10,
                                 leading=14, alignment=2)
    tot_lbl     = ParagraphStyle('TotLbl', parent=styles['Normal'],
                                 fontSize=10, fontName='Helvetica-Bold', alignment=2)
    tot_val     = ParagraphStyle('TotVal', parent=styles['Normal'], fontSize=10, alignment=2)
    grand_lbl   = ParagraphStyle('GrandLbl', parent=styles['Normal'],
                                 fontSize=12, fontName='Helvetica-Bold', alignment=2)
    grand_val   = ParagraphStyle('GrandVal', parent=styles['Normal'],
                                 fontSize=12, fontName='Helvetica-Bold', alignment=2)

    # ── Encabezado ───────────────────────────────────────────────────────
    elements.append(Paragraph('ORDEN DE COMPRA', title_style))
    elements.append(Paragraph('TecnoStock S.A', sub_style))

    # ── Info proveedor / compra ────────────────────────────────────────
    def lp(t): return Paragraph(esc(t), label_style)
    def vp(t): return Paragraph(esc(str(t)), value_style)

    cw = available_width / 4
    info_rows = [
        [lp('Compra #:'),   vp(purchase.id),
         lp('Fecha:'),      vp(purchase.purchase_date.strftime('%d/%m/%Y %H:%M'))],
        [lp('Proveedor:'),  vp(purchase.supplier.name),
         lp('N° Factura:'), vp(purchase.document_number)],
    ]
    if purchase.supplier.email:
        info_rows.append([lp('Email:'), vp(purchase.supplier.email), vp(''), vp('')])
    if purchase.supplier.phone:
        info_rows.append([lp('Teléfono:'), vp(purchase.supplier.phone), vp(''), vp('')])

    info_table = Table(info_rows, colWidths=[cw * 0.8, cw * 1.4, cw * 0.8, cw * 1.0])
    info_table.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND',    (0, 0), (-1, -1), colors.HexColor('#F0F4F8')),
        ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',     (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3 * inch))

    # ── Tabla de detalles ────────────────────────────────────────────────
    col_w = [available_width * p for p in (0.45, 0.15, 0.20, 0.20)]
    detail_rows = [[
        Paragraph('Producto',     hdr_style),
        Paragraph('Cantidad',     hdr_style),
        Paragraph('Costo Unit.',  hdr_style),
        Paragraph('Subtotal',     hdr_style),
    ]]
    for d in purchase.details.all():
        detail_rows.append([
            Paragraph(esc(d.product.name), cell_style),
            Paragraph(str(d.quantity),     cell_r),
            Paragraph(f'${d.unit_cost}',   cell_r),
            Paragraph(f'${d.subtotal}',    cell_r),
        ])

    num_data_rows = len(detail_rows)
    detail_rows.append(['', '', Paragraph('Subtotal:',   tot_lbl), Paragraph(f'${purchase.subtotal}', tot_val)])
    detail_rows.append(['', '', Paragraph('IVA (15%):', tot_lbl), Paragraph(f'${purchase.tax}',      tot_val)])
    detail_rows.append(['', '', Paragraph('TOTAL:',      grand_lbl), Paragraph(f'${purchase.total}', grand_val)])

    det_table = Table(detail_rows, colWidths=col_w, repeatRows=1)
    det_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0),              (-1, 0),                   colors.HexColor('#0F172A')),
        ('LINEBELOW',     (0, 0),              (-1, 0),                   2, colors.HexColor('#0F172A')),
        ('ROWBACKGROUNDS',(0, 1),              (-1, num_data_rows - 1),   [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',          (0, 0),              (-1, num_data_rows - 1),   0.5, colors.HexColor('#CBD5E1')),
        ('LINEABOVE',     (2, num_data_rows),  (3, num_data_rows),        1, colors.HexColor('#0F172A')),
        ('BACKGROUND',    (2, num_data_rows + 2), (3, num_data_rows + 2), colors.HexColor('#fff3cd')),
        ('BOX',           (2, num_data_rows + 2), (3, num_data_rows + 2), 1.5, colors.HexColor('#0F172A')),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    elements.append(det_table)

    # ── Pie ──────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 0.4 * inch))
    foot_style = ParagraphStyle('Foot', parent=styles['Normal'],
                                fontSize=8, textColor=colors.grey, alignment=1)
    elements.append(Paragraph(
        f'Documento generado por TecnoStock S.A. el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}',
        foot_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="compra_{purchase.id}.pdf"'
    return response


@login_required
def purchase_excel(request, pk):
    """Genera y descarga un Excel de la orden de compra."""
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier').prefetch_related('details__product'),
        pk=pk,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Compra #{purchase.id}'

    dark = PatternFill(fill_type='solid', fgColor='0F172A')
    white_bold = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    # Encabezado
    ws.merge_cells('A1:D1')
    ws['A1'] = 'TecnoStock S.A.'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center

    ws.merge_cells('A2:D2')
    ws['A2'] = f'ORDEN DE COMPRA #{purchase.id}'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    # Datos del proveedor
    ws['A4'] = 'Proveedor:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = purchase.supplier.name
    ws['C4'] = 'N° Factura:'
    ws['C4'].font = Font(bold=True)
    ws['D4'] = purchase.document_number

    ws['A5'] = 'Fecha:'
    ws['A5'].font = Font(bold=True)
    ws['B5'] = purchase.purchase_date.strftime('%d/%m/%Y %H:%M')

    # Cabecera de tabla
    for col, header in enumerate(['Producto', 'Cantidad', 'Costo Unitario', 'Subtotal'], 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = white_bold
        cell.fill = dark
        cell.alignment = center

    # Filas de detalle
    row = 8
    for detail in purchase.details.all():
        ws.cell(row=row, column=1, value=detail.product.name)
        ws.cell(row=row, column=2, value=detail.quantity).alignment = center
        ws.cell(row=row, column=3, value=float(detail.unit_cost)).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4, value=float(detail.subtotal)).number_format = '"$"#,##0.00'
        row += 1

    # Totales
    ws.cell(row=row, column=3, value='Subtotal:').font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(purchase.subtotal)).number_format = '"$"#,##0.00'
    row += 1
    ws.cell(row=row, column=3, value='IVA (15%):').font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(purchase.tax)).number_format = '"$"#,##0.00'
    row += 1
    total_cell = ws.cell(row=row, column=3, value='TOTAL:')
    total_cell.font = Font(bold=True, size=12)
    total_val = ws.cell(row=row, column=4, value=float(purchase.total))
    total_val.font = Font(bold=True, size=12)
    total_val.number_format = '"$"#,##0.00'

    # Pie
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    footer = ws[f'A{row}']
    footer.value = f'Documento generado por TecnoStock S.A. — {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    footer.alignment = center
    footer.font = Font(italic=True, color='888888')

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="compra_{purchase.id}.xlsx"'
    return response


@login_required
def purchase_list_excel(request):
    """Exporta la lista de compras filtrada a Excel."""
    purchases = Purchase.objects.select_related('supplier').all()

    supplier_id = request.GET.get('supplier', '')
    if supplier_id:
        purchases = purchases.filter(supplier_id=supplier_id)

    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from and date_to:
        purchases = purchases.filter(purchase_date__date__range=[date_from, date_to])
    elif date_from:
        purchases = purchases.filter(purchase_date__date__gte=date_from)
    elif date_to:
        purchases = purchases.filter(purchase_date__date__lte=date_to)

    year = request.GET.get('year', '')
    if year:
        purchases = purchases.filter(purchase_date__year=year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lista de Compras'

    dark = PatternFill(fill_type='solid', fgColor='0F172A')
    white_bold = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center')

    ws.merge_cells('A1:G1')
    ws['A1'] = 'TecnoStock S.A. — Lista de Compras'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    for col, header in enumerate(['#', 'Proveedor', 'N° Documento', 'Fecha', 'Subtotal', 'IVA', 'Total'], 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = white_bold
        cell.fill = dark
        cell.alignment = center

    for row, purchase in enumerate(purchases, 4):
        ws.cell(row=row, column=1, value=purchase.id).alignment = center
        ws.cell(row=row, column=2, value=purchase.supplier.name)
        ws.cell(row=row, column=3, value=purchase.document_number)
        ws.cell(row=row, column=4, value=purchase.purchase_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=float(purchase.subtotal)).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6, value=float(purchase.tax)).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=7, value=float(purchase.total)).number_format = '"$"#,##0.00'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="compras.xlsx"'
    return response


@login_required
def purchase_report(request):
    """Costo promedio por producto usando aggregate(Avg('unit_cost'))."""
    report = (
        PurchaseDetail.objects
        .values('product__id', 'product__name', 'product__brand__name')
        .annotate(
            avg_cost=Avg('unit_cost'),
            total_quantity=Sum('quantity'),
            total_spent=Sum('subtotal'),
            purchase_count=Count('purchase', distinct=True),
        )
        .order_by('product__name')
    )

    # Totales globales
    totals = PurchaseDetail.objects.aggregate(
        grand_avg=Avg('unit_cost'),
        grand_total=Sum('subtotal'),
        grand_qty=Sum('quantity'),
    )

    return render(request, 'purchasing/purchase_report.html', {
        'report': report,
        'totals': totals,
    })
