from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime
import json
import html

# Importar librerías para exportación
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class StaffRequiredMixin:
    staff_redirect_url = '/'
    staff_error_message = 'No tienes permisos para realizar esta acción. Se requiere acceso de administrador.'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)
        return super().dispatch(request, *args, **kwargs)


class ExportMixin:
    """Mixin genérico para exportar listados a PDF y Excel"""
    export_fields = []  # Debe definirse en la vista: [('field_name', 'Label'), ...]
    export_filename = 'exportar'
    export_title = 'Listado'

    def get_export_data(self):
        """Obtiene los datos del queryset actual con soporte para campos relacionados (brand.name, etc.)"""
        queryset = self.get_queryset()
        data = []

        for obj in queryset:
            row = []
            for field_name, label in self.export_fields:
                try:
                    if '.' in str(field_name):
                        value = obj
                        for part in field_name.split('.'):
                            value = getattr(value, part)
                    else:
                        value = getattr(obj, field_name)

                    if isinstance(value, bool):
                        value = 'Sí' if value else 'No'
                    elif value is None:
                        value = ''
                    else:
                        value = str(value)
                    row.append(value)
                except AttributeError:
                    row.append('')
            data.append(row)

        return data
    
    def export_to_pdf(self):
        """Genera un PDF con los datos"""
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(letter),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30,
            )

            elements = []
            styles = getSampleStyleSheet()

            # Estilos de celda con wrapping habilitado
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=9,
                leading=13,
            )
            header_cell_style = ParagraphStyle(
                'HeaderCellStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.whitesmoke,
                fontName='Helvetica-Bold',
            )

            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=20,
                alignment=1,
            )
            elements.append(Paragraph(self.export_title, title_style))

            # Fecha
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=2,
            )
            elements.append(Paragraph(
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
                date_style,
            ))
            elements.append(Spacer(1, 0.3 * inch))

            # Calcular anchos de columna: descripción recibe más espacio
            page_width = landscape(letter)[0]
            available_width = page_width - 60  # márgenes izq + der

            col_weights = []
            for field_name, label in self.export_fields:
                fname = field_name.lower()
                lname = label.lower()
                if any(k in fname for k in ('description', 'notes', 'address', 'descripci')):
                    col_weights.append(4)
                elif any(k in fname for k in ('name', 'nombre')):
                    col_weights.append(2)
                else:
                    col_weights.append(1)

            total_weight = sum(col_weights) or 1
            col_widths = [w / total_weight * available_width for w in col_weights]

            # Cabeceras como Paragraph
            headers = [
                Paragraph(html.escape(label), header_cell_style)
                for _, label in self.export_fields
            ]

            # Datos como Paragraph (con escape XML para evitar errores)
            raw_data = self.get_export_data()
            table_data = [headers]
            for row in raw_data:
                table_data.append([
                    Paragraph(html.escape(str(cell)), cell_style)
                    for cell in row
                ])

            # Tabla con anchos calculados
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1f4788')),
                ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.whitesmoke),
                ('ALIGN',         (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
                ('TOPPADDING',    (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING',   (0, 0), (-1, -1), 8),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
                ('LINEBELOW',     (0, 0), (-1, 0),  2, colors.HexColor('#1f4788')),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ]))
            elements.append(table)

            # Generar PDF
            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f'attachment; filename="{self.export_filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            )
            return response
        except Exception as e:
            messages.error(self.request, f'Error al generar el PDF: {str(e)}')
            return redirect(self.request.path)
    
    def export_to_excel(self):
        """Genera un Excel con los datos"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.export_title[:31]  # Excel limita a 31 caracteres
            
            # Estilos
            header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [label for field_name, label in self.export_fields]
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            data = self.get_export_data()
            for row_num, row_data in enumerate(data, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Ajustar ancho de columnas
            for col_num in range(1, len(headers) + 1):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = 20
            
            # Generar Excel
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.export_filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
        except Exception as e:
            messages.error(self.request, f'Error al generar el Excel: {str(e)}')
            return redirect(self.request.path)

